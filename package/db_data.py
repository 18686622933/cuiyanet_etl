#!/usr/bin/env python3
# -*- coding:utf-8 -*-

import pymysql
import os
import openpyxl
from package import config_file
from package import execel_data


def get_province(config):
    """
    获取sys_batch中的省份code、name
    :param config: 数据库连接信息
    :return: {省份编码:省份名称...}
    """
    sql_get_province = '''
                        SELECT DISTINCT procode,proname FROM `sys_batch`
                        '''
    conn = pymysql.connect(**config)
    cursor = conn.cursor()
    cursor.execute(sql_get_province)
    provinces = {}
    for i in cursor.fetchall():
        provinces[i[0]] = i[1]

    cursor.close()
    conn.close()
    return provinces


def get_table(config):
    """
    获取库中最新年份的数据对标表名称，并按省份整理，以字典返回
    :param config:
    :return:{'新疆维吾尔自治区': ['entrance_guidefra_650000_2019_298', 'entrance_guidefra_650000_2019_299'],。。。}
    """
    # 获取库中所有数据对标表名称
    sql_get_tables = '''SHOW TABLES'''
    conn = pymysql.connect(**config)
    cursor = conn.cursor()
    cursor.execute(sql_get_tables)
    tables = [i[0] for i in cursor.fetchall() if len(i[0]) > 18 and i[0][:18] == "entrance_guidefra_"]
    cursor.close()
    conn.close()

    # 筛选出年份最新的数据对标表
    max_year = 2018
    tables_dict = {}
    for t in tables:
        if int(t[-8:-4]) > max_year:
            max_year = int(t[-8:-4])

    tables = [i for i in tables if int(i[-8:-4]) == max_year]

    # 对最新年份的数据对标表按省份整理，存放到字典中
    provinces = get_province(config)
    for t in tables:
        pro_name = execel_data.format_pro_name(provinces[t[-15:-9]][:2])
        if pro_name not in tables_dict.keys():
            tables_dict[pro_name] = [t]
        else:
            tables_dict[pro_name].append(t)

    return tables_dict


def file_name(name):
    """
    判断当前目录是否存在指定的文件或文件夹，名称重复则加(n)自增
    :param name: 预计名称
    :return: 完整路径
    """
    target_path = os.listdir()  # 如果想改为在上级目录创建则改为target_path = os.listdir('..')
    if name in target_path:
        fnum = 1
        while True:
            num = '(' + str(fnum) + ')'
            try:
                name_list = list(name)
                idx = name_list.index('.')
                name_list.insert(idx, num)
                new_name = "".join(name_list)
            except:
                new_name = name + num

            if new_name not in target_path:
                name = new_name
                break
            else:
                fnum += 1
    return name


def create_file(pro, year):
    """
    创建主文件夹提供算法_省份(n)，及子文件夹['年份_省份_科类_四年分数','年份_省份_科类_三年分数','年份_省份_科类二年分数']
    :param pro: 省份简称
    :param year: 年份
    :return: /Users/cbowen/PycharmProjects/pymysql/提供算法_省份(n)
    """
    name = file_name('提供算法_%s' % pro)
    os.mkdir('%s' % name)  # 如果想改为在上级目录创建则改为os.mkdir('../%s' % name)
    father = os.path.dirname(os.path.dirname(__file__))
    pro_dir = father + '/%s' % name

    os.mkdir(pro_dir + '/%s_%s_理科_四年分数' % (year, pro))
    os.mkdir(pro_dir + '/%s_%s_理科_三年分数' % (year, pro))
    os.mkdir(pro_dir + '/%s_%s_理科_二年分数' % (year, pro))
    os.mkdir(pro_dir + '/%s_%s_文科_四年分数' % (year, pro))
    os.mkdir(pro_dir + '/%s_%s_文科_三年分数' % (year, pro))
    os.mkdir(pro_dir + '/%s_%s_文科_二年分数' % (year, pro))
    return pro_dir


def get_db_data(config, table_name, directory, pro):
    """
    将整张表的数据放到一个excel中
    :param config: 数据库信息
    :param table_name: 表名称
    :param directory: 文件保存路径
    :return:
    """

    # 将整个表数据保存到excel
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    keys = ['报考书-专业名称', '分数-年份', '最低分', '最高分', '总人数', '最低分排名', '最高分排名', '报考书-ID', '报考书-招生人数', '报考书-学校招生代码',
            '报考书-专业招生代码', '报考书-学校id', '报考书-专业编码', '批次id', '批次名称', '计划人数', '学校报考代码', '专业报考代码', '专业原始名', '学校名称']
    for i in range(1, len(keys) + 1):
        worksheet.cell(row=1, column=i).value = keys[i - 1]
    row = 2

    sql_getdb_data = '''
                    SELECT 
                    enguide_majname,years,minfra,maxfra,sub_sumnums,minrank,maxrank,enguide_id,enguide_num,
                    enguide_sch_bkcode,enguide_maj_bkcode,enguide_schid,enguide_majcode,batch_id,batch_name,
                    enguide_num,enguide_sch_bkcode,enguide_maj_bkcode,enguide_majname,enguide_schname 
                    FROM %s
                    order by years,enguide_sch_bkcode,enguide_maj_bkcode              
                    ''' % table_name
    conn = pymysql.connect(**config)
    cursor = conn.cursor()
    cursor.execute(sql_getdb_data)
    one_row = cursor.fetchone()
    while one_row:
        for i in range(0, len(one_row)):
            worksheet.cell(row=row, column=i + 1).value = one_row[i]
        row += 1
        # print(row, one_row, sep=': ')
        one_row = cursor.fetchone()

    cursor.close()
    # conn.close()
    subject = lambda x: '理科' if x == '298' else '文科'
    workbook.save(directory + '/%s' % subject(table_name[-3:]) + '.xlsx')

    # 将数据按学校和涵盖年份数 拆分
    # 先获取各高校涵盖年份数，放到字典里{1:[...], 2:[...], 3:[...], }
    cursor = conn.cursor()
    years_school = {}
    sql_school_years = '''
                        select enguide_schname,COUNT(*) from 
                        (SELECT years,enguide_schname FROM %s GROUP by years,enguide_schname) a
                        GROUP by enguide_schname HAVING COUNT(*)>=2
                        ''' % table_name

    cursor.execute(sql_school_years)
    one_school = cursor.fetchone()
    while one_school:
        if one_school[1] in years_school.keys():
            years_school[one_school[1]].append(one_school[0])
        else:
            years_school[one_school[1]] = [one_school[0]]
        one_school = cursor.fetchone()
    cursor.close()
    # conn.close()

    # 对保存年份数及学校名称的dict进行遍历，并按校名查询数据放到指定的excel中，
    for nums, schools in years_school.items():
        for school_name in schools:
            split_data(table_name, school_name, keys, conn, directory, nums, pro)

    conn.close()


def split_data(table_name, school_name, keys, conn, directory, nums, pro):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    for i in range(1, len(keys)):
        worksheet.cell(row=1, column=i).value = keys[i - 1]
    row = 2
    sql_getdb_data = '''
                    SELECT 
                    enguide_majname,years,minfra,maxfra,sub_sumnums,minrank,maxrank,enguide_id,enguide_num,
                    enguide_sch_bkcode,enguide_maj_bkcode,enguide_schid,enguide_majcode,batch_id,batch_name,
                    enguide_num,enguide_sch_bkcode,enguide_maj_bkcode,enguide_majname
                    FROM %s where enguide_schname = '%s'
                    order by years,enguide_sch_bkcode,enguide_maj_bkcode              
                    ''' % (table_name, school_name)
    cursor = conn.cursor()
    cursor.execute(sql_getdb_data)
    one_row = cursor.fetchone()
    while one_row:
        for i in range(0, len(one_row) - 1):
            worksheet.cell(row=row, column=i + 1).value = one_row[i]
        row += 1
        # print(row, one_row, sep=': ')
        one_row = cursor.fetchone()

    cursor.close()
    # conn.close()

    workbook.save(directory + '/%s' % sub_dir(table_name, nums, pro) + '/%s.xlsx' % school_name)


def year_nums_to_chinese(number):
    num_dict = {0: "零", 1: "一", 2: "二", 3: "三", 4: "四", 5: "五", 6: "六", 7: "七", 8: "八", 9: "九"}
    return num_dict[number]


def sub_dir(table_name, nums, pro):
    year = table_name[-8:-4]
    province = pro
    to_subject = lambda x: '理科' if x == '298' else '文科'
    subject = to_subject(table_name[-3:])
    nums_chinese = "%s年分数" % year_nums_to_chinese(nums)

    result = "_".join([year, pro, subject, nums_chinese])
    return result

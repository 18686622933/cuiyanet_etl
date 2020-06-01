#!/usr/bin/env python3
# -*- coding:utf-8 -*-

import os
import openpyxl
import re
import pymysql


def format_pro_name(pro_name):
    if pro_name == '黑龙':
        return '黑龙江'
    elif pro_name == '内蒙':
        return '内蒙古'
    else:
        return pro_name


def parser_file_name(file_name):
    result = {'province_sort': ' ', 'year': ' ', 'key': ' '}
    try:
        result['province_sort'] = format_pro_name(file_name[:2])
    except:
        pass
    try:
        result['year'] = re.findall("[0-9]{4}", file_name)[0]
    except:
        pass
    try:
        result['key'] = re.findall("报考书|拆分表|录取分数|数据对标|一分一段表", file_name)[0]
    except:
        pass
    return result


def parser_dir(dir='.', keys={'bks': '报考书', 'cfb': '拆分表', 'lqfs': '录取分数', 'sjdb': '数据对标', 'yfydb': '一分一段表'}):
    """
    将当前目录下文件按关键字分组，放到字典gruop_names中
    *要求：文件名必须由省份开头
    :param dir: 默认为当前目录
    :param keys: 指定文件名中的关键字
    :return: group_names
    {'bks': ['新疆报考书入库.xlsx'], 'cfb': ['新疆2019年拆分表.xlsx'], 'lqfs': ['黑龙江录取分数2018的副本.xlsx', '新疆录取分数2018.xlsx'], 'sjdb': ['黑龙江数据对标历年分数最终版的副本.xlsx', '新疆数据对标历年分数最终版.xlsx'], 'yfydb': ['吉林一分一段表的副本.xlsx', '新疆一分一段表.xlsx']}

    """
    dirlist = os.listdir(dir)
    # print(dirlist)

    gruop_names = {}
    for v in keys.values():
        gruop_names[v] = []

    for name in dirlist:
        if name[-5:] == '.xlsx' and name[:2] != '~$':
            for k in keys:
                if re.search(keys[k], name):
                    gruop_names[keys[k]].append(name)
    # print(gruop_names)
    return gruop_names


def to_distribution(file_name, config, province_data, type={'综合': 407, '理科': 298, '文科': 299}):
    """
    一分一段表导入
    根据文件名中的省份，判断sys_distribution是否存在同省份数据，存在则提示请核对
    否则将excel中数据关联基础数据（省份信息、科类信息）传入sys_distribution中
    :param file_name: 数据来源excel名称
    :param config: 数据库信息
    :param province_data: 省份信息
    :param type: 科类信息
    :return: 传输结果
    """

    province_sort = file_name[:2]  # 在文件名中获取省份
    conn = pymysql.connect(**config)
    cursor = conn.cursor()

    # 检查目标表中是否存该省份数据
    sql_province_exist = '''
            select id from  sys_distribution where substr(province,1,2) = %s 
                    '''
    cursor.execute(sql_province_exist, province_sort)
    province_exist = cursor.fetchone()

    # 存在则提示请核对
    if province_exist:
        cursor.close()
        conn.close()
        print("！！！ %s%s%s 数据已存在，请核对!" % (parser_file_name(file_name)['province_sort'],
                                     parser_file_name(file_name)['year'], parser_file_name(file_name)['key']))

    # 不存在则进行数据上传
    else:
        wb = openpyxl.load_workbook(file_name)  # 打开excel
        sheet = wb.worksheets[0]  # 获取第一个表
        max_row = sheet.max_row
        max_column = sheet.max_column

        column_excel = [sheet.cell(row=1, column=i).value.lower() for i in range(1, max_column + 1)]
        # print(column_excel)  # excel所有字段名称

        order_reference = {}
        for column in column_excel:
            if re.search('年份', column):
                order_reference['年份'] = column_excel.index(column) + 1
            elif re.search('科类', column):
                order_reference['科类'] = column_excel.index(column) + 1
            elif re.search('位次', column):
                order_reference['位次'] = column_excel.index(column) + 1
            elif re.search('最低分', column):
                order_reference['最低分'] = column_excel.index(column) + 1
        # print(order_reference)  # 对字段名称进行正则匹配，用字典保存列数，以便后面插入数据时匹配顺序

        sql_insert = '''
            insert into sys_distribution(year,type_name,num,score,province,pro_id,type) values(%s,%s,%s,%s,%s,%s,%s)
                    '''
        print("--- 正在导入 --- %s%s%s 数据，请等待。。。" % (parser_file_name(file_name)['province_sort'],
                                        parser_file_name(file_name)['year'], parser_file_name(file_name)['key']))
        num = 0
        for i in range(2, max_row + 1):
            one_row = []
            one_row.append(sheet.cell(row=i, column=order_reference['年份']).value)
            one_row.append(sheet.cell(row=i, column=order_reference['科类']).value)
            one_row.append(sheet.cell(row=i, column=order_reference['位次']).value)
            one_row.append(sheet.cell(row=i, column=order_reference['最低分']).value)
            one_row.append(province_sort)
            one_row.append(province_data[province_sort]['id'])
            one_row.append(type[sheet.cell(row=i, column=order_reference['科类']).value])
            cursor.execute(sql_insert, one_row)
            num += 1
            # print(i, one_row, sep=':')

        conn.commit()
        cursor.close()
        conn.close()
        # print(format_pro_name(province_sort) + "数据上传完毕，共计%s条，请检查！" % num)
        print("vvv 导入成功 vvv %s%s%s 共计上传%s条，请检查。" % (parser_file_name(file_name)['province_sort'],
                                             parser_file_name(file_name)['year'], parser_file_name(file_name)['key'],
                                             num))


def to_fractional(file_name, config, sch_data, major_data, batch_data, subject={'综合': 407, '理科': 298, '文科': 299}):
    """
    录取分数导入
    判断库里是否存在同年份同省份的数据，有则提示请核对，没有则上传。
    :param file_name: 文件名
    :param config: 数据库信息
    :param sch_data: 学校数据
    :param major_data: 专业数据
    :param batch_data: 批次数据
    :param subject: 科类数据
    :return: 上传结果
    """
    province_sort = file_name[:2]  # 在文件名中获取省份
    year = re.findall("[0-9]{4}", file_name)

    conn = pymysql.connect(**config)
    cursor = conn.cursor()

    # 检查目标表中是否存省份和年份相同的数据
    sql_province_exist = '''
            select id from hm_school_major_fractional where substr(pro_name,1,2) = %s and years = %s
                        '''
    cursor.execute(sql_province_exist, (province_sort, year))
    province_exist = cursor.fetchone()

    # 存在则提示请核对
    if province_exist:
        cursor.close()
        conn.close()
        print("！！！ %s%s%s 数据已存在，请核对!" % (parser_file_name(file_name)['province_sort'],
                                     parser_file_name(file_name)['year'], parser_file_name(file_name)['key']))

    # 不存在则进行数据上传
    else:
        wb = openpyxl.load_workbook(file_name)  # 打开excel
        sheet = wb.worksheets[0]  # 获取第一个表
        max_row = sheet.max_row
        max_column = sheet.max_column

        column_excel = [sheet.cell(row=1, column=i).value.lower() for i in range(1, max_column + 1)]
        # print(column_excel)  # excel所有字段名称

        order_reference = {'school_name': '学校名称', 'old_major_name': '专业名称', 'major_name': '专业名称去括号', 'years': '年份',
                           'pro_id': 'id', 'pro_name': '省份', 'subject_name': '科类', 'batch_name': '批次', 'minfrac': '最低分',
                           'initnums': '录取人数', 'maxfrac': '最高分', 'avgfrac': '平均分', 'minrank': '最低分位次'}
        sql_insert = '''
        insert into
        hm_school_major_fractional(school_name,old_major_name,major_name,years,
                         pro_id,pro_name,subject_name,batch_name,minfrac,initnums,maxfrac,avgfrac,minrank,
                         school_id,school_code,major_id,major_code,subject_id,batch_id)
                         values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    '''

        print("--- 正在导入 --- %s%s%s 数据，请等待。。。" % (parser_file_name(file_name)['province_sort'],
                                        parser_file_name(file_name)['year'], parser_file_name(file_name)['key']))
        num = 0
        unmatch = 0
        unmatch_school = set()
        unmatch_major = set()
        for i in range(2, max_row + 1):
            one_row = []
            for k in order_reference.keys():
                one_row.append(sheet.cell(row=i, column=column_excel.index(order_reference[k]) + 1).value)

            # # 在sch_data中按学校名称匹配
            # if sheet.cell(row=i,
            #               column=column_excel.index(order_reference['school_name']) + 1).value in sch_data.keys():
            #     one_row.append(
            #         sch_data[sheet.cell(row=i, column=column_excel.index(order_reference['school_name']) + 1).value][
            #             'id'])
            #     one_row.append(
            #         sch_data[sheet.cell(row=i, column=column_excel.index(order_reference['school_name']) + 1).value][
            #             'code'])
            # # 如果匹配不到则通过正则进行匹配
            # else:
            #     got_same = 0
            #     for k in sch_data:
            #         re_school = re.search(
            #             sheet.cell(row=i, column=column_excel.index(order_reference['school_name']) + 1).value, k)
            #         if re_school:
            #             got_same = 1
            #             one_row.append(sch_data[k]['id'])
            #             one_row.append(sch_data[k]['code'])
            #             break
            #     # 如果正则也匹配不到则将school_id=0，school_code=000
            #     if got_same == 0:
            #         one_row.append(0)
            #         one_row.append(000)

            try:
                one_row.append(
                    sch_data[
                        sheet.cell(row=i, column=column_excel.index(order_reference['school_name']) + 1).value]['id'])
            except:
                one_row.append(None)
                unmatch_school.add(
                    sheet.cell(row=i, column=column_excel.index(order_reference['school_name']) + 1).value)

            try:
                one_row.append(
                    sch_data[
                        sheet.cell(row=i, column=column_excel.index(order_reference['school_name']) + 1).value]['id'])
            except:
                one_row.append(None)
                unmatch_school.add(
                    sheet.cell(row=i, column=column_excel.index(order_reference['school_name']) + 1).value)

            try:
                one_row.append(
                    major_data[sheet.cell(row=i, column=column_excel.index(order_reference['major_name']) + 1).value][
                        'id'])
            except:
                one_row.append(None)
                unmatch_major.add(sheet.cell(row=i, column=column_excel.index(order_reference['major_name']) + 1).value)

            try:
                one_row.append(
                    major_data[sheet.cell(row=i, column=column_excel.index(order_reference['major_name']) + 1).value][
                        'code'])
            except:
                one_row.append(None)
                unmatch_major.add(sheet.cell(row=i, column=column_excel.index(order_reference['major_name']) + 1).value)

            one_row.append(
                subject[sheet.cell(row=i, column=column_excel.index(order_reference['subject_name']) + 1).value])

            # 按一、二、三匹配批次
            for k in batch_data:
                if k[:2] == province_sort and re.search("[一二三]", k).group() == re.search(
                        "[一二三]",
                        sheet.cell(row=i, column=column_excel.index(order_reference['batch_name']) + 1).value).group():
                    one_row.append(batch_data[k]['id'])

            for v in one_row[14:18]:
                if v is None:
                    unmatch = 1
                    # print("问题数据：", i, one_row, sep=' ')
                    break

            # 导入数据库，不成功则打印
            try:
                cursor.execute(sql_insert, (
                    one_row[0], one_row[1], one_row[2], one_row[3], one_row[4], one_row[5], one_row[6], one_row[7],
                    one_row[8], one_row[9], one_row[10], one_row[11], one_row[12], one_row[13], one_row[14],
                    one_row[15], one_row[16], one_row[17], one_row[18]))
                num += 1
            except:
                print("xxx 导入失败 xxx：", one_row, sep=' ')

        cursor.close()
        if unmatch == 1:
            unmatch_data = {}
            unmatch_data['学校'] = unmatch_school
            unmatch_data['专业'] = unmatch_major
            for k in list(unmatch_data.keys()):
                if unmatch_data[k] == set():
                    del unmatch_data[k]
            print("xxx 导入失败 xxx 基础表中存在缺失项：", unmatch_data, sep='')
        else:
            conn.commit()
            print("vvv 导入成功 vvv %s%s%s 共计上传%s条，请检查。" % (
                parser_file_name(file_name)['province_sort'], parser_file_name(file_name)['year'],
                parser_file_name(file_name)['key'], num))
        conn.close()


def to_guidefra(file_name, config, province_data, sch_data, major_data, batch_data,
                subject={'综合': 407, '理科': 298, '文科': 299}):
    """
    报考书导入
    :param file_name:
    :param config:
    :param sch_data:
    :param major_data:
    :param batch_data:
    :param subject:
    :return:
    """
    province_sort = file_name[:2]  # 在文件名中获取省份
    province_code = province_data[format_pro_name(province_sort)]['provincecode']
    year = re.findall("[0-9]{4}", file_name)[0]
    table_name = 'entrance_guidelines_{0}_{1}'.format(province_code, year)
    # table_name = 'entrance_guidelines_650000_2018'
    conn = pymysql.connect(**config)
    cursor = conn.cursor()

    do_insert = 0
    # 查看表名是否存在,不存在则创建，存在则判断是否为空，新建表或表中数据为空则允许插入do_insert = 1，否则提示数据已存在
    sql_table_exist = '''show tables like \'%s\'''' % table_name
    cursor.execute(sql_table_exist)
    if cursor.fetchone() is None:
        sql_create_table = '''
            CREATE TABLE %s (
              `id` int(10) NOT NULL AUTO_INCREMENT,
              `sm_id` int(10) DEFAULT NULL COMMENT '学校开设专业ID',
              `school_entrance_code` varchar(20) CHARACTER SET utf8mb4 NOT NULL COMMENT '学校招生代码',
              `school_id` int(10) DEFAULT NULL COMMENT '学校id',
              `school_name` varchar(20) CHARACTER SET utf8mb4 NOT NULL COMMENT '学校名称',
              `major_entrance_code` varchar(20) CHARACTER SET utf8mb4 NOT NULL COMMENT '专业招生代码',
              `major_id` int(10) DEFAULT NULL COMMENT '专业id(弃用)',
              `major_code` varchar(20) DEFAULT NULL COMMENT '专业或者一级学科（自定义编码）',
              `old_major_name` text COMMENT '专业名称（报考书原始名）匹配分数使用',
              `major_name` varchar(120) CHARACTER SET utf8mb4 DEFAULT NULL COMMENT '专业名称（去括号拆后）网站展示-关联hm_major使用',
              `sm_major_name` varchar(200) DEFAULT NULL COMMENT '学校专业名（处理后）关联hm_school_major使用',
              `major_sm` text COMMENT '专业说明',
              `major_yq` text COMMENT '招考要求',
              `edu_year` varchar(20) CHARACTER SET utf8mb4 DEFAULT NULL COMMENT '学制（年）',
              `plan_nums` int(6) DEFAULT NULL COMMENT '计划数（招生人数）',
              `cose_year` varchar(20) CHARACTER SET utf8mb4 DEFAULT NULL COMMENT '年收费（元）',
              `batch_id` int(10) DEFAULT NULL COMMENT '批次id',
              `batch_name` varchar(20) CHARACTER SET utf8mb4 NOT NULL COMMENT '批次名称',
              `subjects_id` int(10) DEFAULT NULL COMMENT '考生类别id',
              `subjects_name` varchar(10) CHARACTER SET utf8mb4 NOT NULL COMMENT '考生类别名称',
              `physics_course` int(1) DEFAULT '0' COMMENT '是否选择物理科目（1，选择，2，不选择）',
              `chemical_course` int(1) DEFAULT '0' COMMENT '是否选择化学科目（1，选择，2，不选择）',
              `biological_course` int(1) DEFAULT '0' COMMENT '是否选择生物科目（1，选择，2，不选择）',
              `geographic_course` int(1) DEFAULT '0' COMMENT '是否选择地理科目（1，选择，2，不选择）',
              `history_course` int(1) DEFAULT '0' COMMENT '是否选择历史科目（1，选择，2，不选择）',
              `political_course` int(1) DEFAULT '0' COMMENT '是否选择政治科目（1，选择，2，不选择）',
              `technology_course` int(1) DEFAULT '0' COMMENT '是否选择技术科目（1，选择，2，不选择）',
              `choose_course` varchar(20) CHARACTER SET utf8mb4 DEFAULT NULL COMMENT '选考科目',
              `page_num` int(10) NOT NULL COMMENT '报考书页码',
              `excstate` int(1) DEFAULT NULL COMMENT '报考书历年分数，分数区分标识（\r\n1 （19,18,17,16，3，4年都有数据的有概率，其他情况都没有概率）\r\n2 （19,18,17,16，没有数据，15年有数据的，没有概率展示推荐，时隔多年再次招生）\r\n3 （19,18,17,16，只有一年或者2年，用最近一年的最低分和当年分数进行上下浮动30分进行比较，在范围内进行推荐，否则不推荐）\r\n4 （19,18,17,16,15年等所有年都没有数据的，新设专业推荐，无历年数据） ）',
              `majstate` int(1) DEFAULT NULL COMMENT '1、只有18年\r\n2、18、17年分数\r\n3、三年分数\r\n4、四年分数\r\n5、没有分数',
              `isexcflag` int(1) DEFAULT NULL COMMENT '数据异常标识（1 正常，2 异常，3已处理）',
              `pre_major` int(11) DEFAULT NULL COMMENT '前一年报考书id',
              `flag` int(11) DEFAULT NULL COMMENT '是否经过分词处理  1 处理过  2未处理  ',
              `majstate_new` varchar(200) DEFAULT NULL COMMENT '1、有四年分数\r\n2、有三年分数\r\n3、有17、18俩年分数\r\n4、没有分数 \r\n5、只有一年分数 \r\n6、不是17、18的俩年分数\r\n（1、4、5、6状态数据正确，2、3不正确，2的状态有一些被3更新了）',
              PRIMARY KEY (`id`) USING BTREE,
              KEY `school_id` (`school_id`) USING BTREE,
              KEY `major_id` (`major_id`) USING BTREE,
              KEY `batch_id` (`batch_id`) USING BTREE,
              KEY `id` (`id`) USING BTREE,
              KEY `major_entrance_code` (`major_entrance_code`) USING BTREE,
              KEY `subjects_id` (`subjects_id`) USING BTREE
            ) ENGINE=InnoDB AUTO_INCREMENT=10979 DEFAULT CHARSET=utf8 COMMENT='报考书+省份code+年 （报考书信息）'
                            ''' % table_name

        try:
            cursor.execute(sql_create_table)
        except:
            print("创建表 \'" + table_name + "\' 失败")
            cursor.close()
            conn.commit()
            conn.close()
        # print("表 \'" + table_name + "\' 已创建")
        do_insert = 1

    else:
        sql_table_count = '''select count(*) from %s''' % table_name
        cursor.execute(sql_table_count)
        count = cursor.fetchone()[0]
        if count >= 1:
            cursor.close()
            conn.close()
            print("！！！ %s%s%s 数据已存在，请核对!" % (parser_file_name(file_name)['province_sort'],
                                         parser_file_name(file_name)['year'], parser_file_name(file_name)['key']))
        else:
            do_insert = 1

    if do_insert == 1:
        wb = openpyxl.load_workbook(file_name)  # 打开excel
        sheet = wb.worksheets[0]  # 获取第一个表
        max_row = sheet.max_row
        max_column = sheet.max_column

        column_excel = [sheet.cell(row=1, column=i).value.lower() for i in range(1, max_column + 1) if
                        sheet.cell(row=1, column=i).value is not None]
        # print(column_excel)  # excel所有字段名称

        order_reference = {'school_entrance_code': '院校代号', 'id': 'id', 'school_name': '学校名称',
                           'major_entrance_code': '专业代号', 'old_major_name': '专业名称合并', 'major_name': '专业名称去括号',
                           'edu_years': '学制', 'plan_nums': '招生计划', 'cose_year': '学费标准', 'batch_name': '批次',
                           'subject_name': '科类', 'page_num': '页码', 'majstate': '历年分数区分标识', 'majstate_new': '历年分数区分标识'}
        sql_insert = '''
        insert into {}(school_entrance_code,id,school_name,major_entrance_code,old_major_name,major_name,
                        edu_year,plan_nums,cose_year,batch_name,subjects_name,page_num,majstate,majstate_new,
                        school_id,major_id,major_code,batch_id,subjects_id)
                        values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                       '''.format(table_name)

        print("--- 正在导入 --- %s%s%s 数据，请等待。。。" % (parser_file_name(file_name)['province_sort'],
                                        parser_file_name(file_name)['year'], parser_file_name(file_name)['key']))

        num = 0
        unmatch = 0
        unmatch_school = set()
        unmatch_major = set()
        for i in range(2, max_row + 1):
            one_row = []
            if sheet.cell(row=i, column=1).value is not None:
                for k in order_reference.keys():
                    one_row.append(sheet.cell(row=i, column=column_excel.index(order_reference[k]) + 1).value)

                try:
                    one_row.append(sch_data[sheet.cell(row=i, column=column_excel.index(
                        order_reference['school_name']) + 1).value]['id'])
                except:
                    one_row.append(None)
                    unmatch_school.add(
                        sheet.cell(row=i, column=column_excel.index(order_reference['school_name']) + 1).value)

                try:
                    one_row.append(
                        major_data[
                            sheet.cell(row=i, column=column_excel.index(order_reference['major_name']) + 1).value][
                            'id'])
                except:
                    one_row.append(None)
                    unmatch_major.add(
                        sheet.cell(row=i, column=column_excel.index(order_reference['major_name']) + 1).value)

                try:
                    one_row.append(major_data[sheet.cell(row=i, column=column_excel.index(
                        order_reference['major_name']) + 1).value]['code'])
                except:
                    one_row.append(None)

                # 按一、二、三匹配批次
                for k in batch_data:
                    if k[:2] == province_sort and re.search("[一二三]", k).group() == re.search(
                            "[一二三]",
                            sheet.cell(row=i,
                                       column=column_excel.index(order_reference['batch_name']) + 1).value).group():
                        one_row.append(batch_data[k]['id'])

                one_row.append(
                    subject[sheet.cell(row=i, column=column_excel.index(order_reference['subject_name']) + 1).value])

            else:
                continue

            for v in one_row[15:18]:
                if v is None:
                    unmatch = 1
                    # print("问题数据：", i, one_row, sep=' ')
                    break

            # 导入数据库，不成功则打印
            try:
                cursor.execute(sql_insert, (
                    one_row[0], one_row[1], one_row[2], one_row[3], one_row[4], one_row[5], one_row[6],
                    one_row[7], one_row[8], one_row[9], one_row[10], one_row[11], one_row[12], one_row[13], one_row[14],
                    one_row[15], one_row[16], one_row[17], one_row[18]))
                num += 1
            except:
                print("xxx 导入失败 xxx：", one_row, sep=' ')

        cursor.close()
        if unmatch == 1:
            unmatch_data = {}
            unmatch_data['学校'] = unmatch_school
            unmatch_data['专业'] = unmatch_major
            for k in list(unmatch_data.keys()):
                if unmatch_data[k] == set():
                    del unmatch_data[k]
            print("xxx 导入失败 xxx 基础表中存在缺失项：", unmatch_data, sep='')
        else:
            conn.commit()
            print("vvv 导入成功 vvv %s%s%s 共计上传%s条，请检查。" % (
                parser_file_name(file_name)['province_sort'], parser_file_name(file_name)['year'],
                parser_file_name(file_name)['key'], num))
        conn.close()


def to_frac(file_name, config, sch_data, major_data, batch_data, category_data, small_class_data,
            subject={'综合': 407, '理科': 298, '文科': 299}):
    province_sort = file_name[:2]  # 在文件名中获取省份
    year = re.findall("[0-9]{4}", file_name)

    conn = pymysql.connect(**config)
    cursor = conn.cursor()

    # 检查目标表中是否存省份和年份相同的数据
    sql_province_exist = '''
            select id from hm_school_major_frac where substr(pro_name,1,2) = %s and years = %s
                        '''
    cursor.execute(sql_province_exist, (province_sort, year))
    province_exist = cursor.fetchone()

    # 存在则提示请核对
    if province_exist:
        cursor.close()
        conn.close()
        print("！！！%s%s%s 数据已存在，请核对!" % (parser_file_name(file_name)['province_sort'],
                                     parser_file_name(file_name)['year'], parser_file_name(file_name)['key']))

    # 不存在则进行数据上传
    else:
        wb = openpyxl.load_workbook(file_name)  # 打开excel
        sheet = wb.worksheets[0]  # 获取第一个表
        max_row = sheet.max_row
        max_column = sheet.max_column

        column_excel = [sheet.cell(row=1, column=i).value.lower() for i in range(1, max_column + 1)]
        # print(column_excel)  # excel所有字段名称

        order_reference = {'school_name': '学校名称', 'major_name': '开设专业', 'years': '年份',
                           'pro_id': '省份id', 'pro_name': '省份', 'subject_name': '科类', 'batch_name': '批次',
                           'minfrac': '最低分', 'plan_nums': '招生人数', 'strong_subject': '优势学科', 'unique_subject': '特色学科',
                           'estimate': '评估结果', 'remark': '备注', 'major_c_name': '招生专业', 'category': '学科门类',
                           'small_class': '一级学科', 'number': '招生人数', 'enguide_id': '报考书id'}
        match_key = ['school_id', 'major_id', 'subject_id', 'batch_id', 'category_id', 'small_class_id']
        sql_insert = '''
        insert into
        hm_school_major_frac(school_name,major_name,years,pro_id,pro_name,subject_name,batch_name,
                                   minfrac,plan_nums,strong_subject,unique_subject,
                                   estimate,remark,major_c_name,category,small_class,number,enguide_id,
                                   school_id,major_id,subject_id,batch_id,category_id,small_class_id)
                         values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    '''

        print("--- 正在导入 --- %s%s%s 数据，请等待。。。" % (parser_file_name(file_name)['province_sort'],
                                        parser_file_name(file_name)['year'], parser_file_name(file_name)['key']))

        num = 0
        unmatch = 0
        unmatch_school = set()
        unmatch_major = set()
        unmatch_categorg = set()
        unmatch_samll_class = set()
        for i in range(2, max_row + 1):
            one_row = []
            if sheet.cell(row=i, column=1).value is not None:
                for k in order_reference.keys():
                    one_row.append(sheet.cell(row=i, column=column_excel.index(order_reference[k]) + 1).value)

                try:
                    one_row.append(
                        sch_data[
                            sheet.cell(row=i, column=column_excel.index(order_reference['school_name']) + 1).value][
                            'id'])
                except:
                    one_row.append(None)
                    unmatch_school.add(
                        sheet.cell(row=i, column=column_excel.index(order_reference['school_name']) + 1).value)

                try:
                    one_row.append(
                        major_data[
                            sheet.cell(row=i, column=column_excel.index(order_reference['major_name']) + 1).value][
                            'id'])
                except:
                    one_row.append(None)
                    unmatch_major.add(
                        sheet.cell(row=i, column=column_excel.index(order_reference['major_name']) + 1).value)

                one_row.append(
                    subject[sheet.cell(row=i, column=column_excel.index(order_reference['subject_name']) + 1).value])

                # 按一、二、三匹配批次
                for k in batch_data:
                    if k[:2] == province_sort and re.search("[一二三]", k).group() == re.search(
                            "[一二三]",
                            sheet.cell(row=i,
                                       column=column_excel.index(order_reference['batch_name']) + 1).value).group():
                        one_row.append(batch_data[k]['id'])

                try:
                    one_row.append(
                        category_data[
                            sheet.cell(row=i, column=column_excel.index(order_reference['category']) + 1).value])
                except:
                    one_row.append(None)
                    unmatch_categorg.add(
                        sheet.cell(row=i, column=column_excel.index(order_reference['category']) + 1).value)

                try:
                    one_row.append(small_class_data[
                                       sheet.cell(row=i,
                                                  column=column_excel.index(order_reference['small_class']) + 1).value])
                except:
                    one_row.append(None)
                    unmatch_samll_class.add(
                        sheet.cell(row=i, column=column_excel.index(order_reference['small_class']) + 1).value)

            else:
                continue

            for v in one_row[19:]:
                if v is None:
                    unmatch = 1
                    # print("问题数据：", i, one_row, sep=' ')
                    break

            # 导入数据库，不成功则打印
            try:
                cursor.execute(sql_insert, (
                    one_row[0], one_row[1], one_row[2], one_row[3], one_row[4], one_row[5], one_row[6],
                    one_row[7], one_row[8], one_row[9], one_row[10], one_row[11], one_row[12], one_row[13], one_row[14],
                    one_row[15], one_row[16], one_row[17], one_row[18], one_row[19], one_row[20], one_row[21],
                    one_row[22], one_row[23]))
                num += 1
            except:
                print("xxx 导入失败 xxx：", one_row, sep=' ')

        cursor.close()
        if unmatch == 1:
            unmatch_data = {}
            unmatch_data['学校'] = unmatch_school
            unmatch_data['专业'] = unmatch_major
            unmatch_data['学科门类'] = unmatch_categorg
            unmatch_data['一级学科'] = unmatch_samll_class
            for k in list(unmatch_data.keys()):
                if unmatch_data[k] == set():
                    del unmatch_data[k]
            print("xxx 导入失败 xxx 基础表中存在缺失项：", unmatch_data, sep='')
        else:
            conn.commit()
            print("vvv 导入成功 vvv %s%s%s 共计上传%s条，请检查。" % (
                parser_file_name(file_name)['province_sort'], parser_file_name(file_name)['year'],
                parser_file_name(file_name)['key'], num))
        conn.close()


def to_guidefra_arts_science(file_name, config, province_data, sch_data, major_data, batch_data,
                             subject={'综合': 407, '理科': 298, '文科': 299}):
    province_sort = file_name[:2]  # 在文件名中获取省份
    province_code = province_data[format_pro_name(province_sort)]['provincecode']
    year = re.findall("[0-9]{4}", file_name)[0]
    table_name_8 = 'entrance_guidefra_{0}_{1}_298'.format(province_code, year)
    table_name_9 = 'entrance_guidefra_{0}_{1}_299'.format(province_code, year)
    # table_name = 'entrance_guidelines_650000_2018'
    # 查看表名是否存在,不存在则创建，存在则判断是否为空，新建表或表中数据为空则允许插入do_insert = 1，否则提示数据已存在
    sql_create_table8 = '''
        CREATE TABLE %s(
          `id` int(10) NOT NULL AUTO_INCREMENT,
          `enguide_id` int(10) NOT NULL COMMENT '报考书-id',
          `enguide_sch_bkcode` varchar(20) NOT NULL COMMENT '报考书-学校招生代码',
          `enguide_maj_bkcode` varchar(20) NOT NULL COMMENT '报考书-专业招生代码',
          `enguide_schid` int(10) DEFAULT NULL COMMENT '报考书-学校id',
          `enguide_schname` varchar(20) NOT NULL COMMENT '报考书-学校名称',
          `enguide_majid` int(10) DEFAULT NULL COMMENT '报考书-专业id（弃用）',
          `enguide_majcode` varchar(20) DEFAULT NULL COMMENT '报考书-专业或者一级学科（萃言自定义编码）',
          `enguide_majname` text NOT NULL COMMENT '报考书-专业名称（报考书原始名）',
          `enguide_subjectid` int(10) DEFAULT NULL COMMENT '报考书-科类编码  299,298,300',
          `enguide_subjectname` varchar(10) DEFAULT NULL COMMENT '报考书_科类名称',
          `sm_id` int(10) DEFAULT NULL COMMENT '报考书-学校专业id',
          `hmschmajfra_id` int(10) DEFAULT NULL COMMENT '历年分数-id（当多个子类分数时id不填）',
          `major` text COMMENT '历年分数-专业名称（分数表专业原始名称）（当多个子类分数时名称用名称|拼接）',
          `subjectid` int(10) DEFAULT NULL COMMENT '历年分数_科类编码 299,298,300',
          `subjectname` varchar(10) NOT NULL COMMENT '历年分数_科类名称',
          `years` varchar(10) NOT NULL COMMENT '历年分数-年份',
          `maxfra` double(6,1) DEFAULT NULL COMMENT '历年分数-最高分',
          `minfra` double(6,1) DEFAULT NULL COMMENT '历年分数-最低分',
          `avgfra` double(6,1) DEFAULT NULL COMMENT '历年分数-平均分',
          `enguide_majstate` int(2) DEFAULT NULL COMMENT '报考书历年分数，专业名称标识（\r\n1.名称一对一\r\n2.大小类转换（跟到学校）（书名无括号）\r\n3.ES查询\r\n4. 分名含于书名（精准）（分数有括号：包含多括号、内嵌括号）\r\n5.中外合作办学\r\n6.只招少数民族\r\n10.分名含于书名（精准）（分名无括号、书名单括号（包含内嵌括号））\r\n11\r\n12\r\n',
          `sub_sumnums` int(8) DEFAULT NULL COMMENT '历年分数-科类下考生总人数',
          `minrank` int(8) DEFAULT NULL COMMENT '历年分数-最低分当年排名',
          `maxrank` int(8) DEFAULT NULL COMMENT '历年分数-最高分当年排名',
          `avgrank` int(8) DEFAULT NULL COMMENT '历年分数-平均分当年排名（暂时弃用）（程序暂时未添加）',
          `state` int(1) DEFAULT NULL COMMENT '数据异常状态',
          `batch_id` int(10) DEFAULT NULL COMMENT '历年分数-批次id',
          `batch_name` varchar(20) DEFAULT NULL COMMENT '历年分数-批次名称',
          `enguide_num` int(8) DEFAULT NULL COMMENT '报考书招生人数',
          `enguide_batch_id` int(10) DEFAULT NULL COMMENT '报考书_历年分数-批次id',
          `enguide_batch_name` varchar(20) DEFAULT NULL COMMENT '报考书_历年分数-批次名称',
          PRIMARY KEY (`id`) USING BTREE,
          KEY `index_1` (`enguide_id`) USING BTREE
        ) ENGINE=InnoDB AUTO_INCREMENT=14552 DEFAULT CHARSET=utf8mb4 COMMENT='报考书与历年分数匹配'
                        ''' % table_name_8
    sql_create_table9 = '''
            CREATE TABLE %s(
              `id` int(10) NOT NULL AUTO_INCREMENT,
              `enguide_id` int(10) NOT NULL COMMENT '报考书-id',
              `enguide_sch_bkcode` varchar(20) NOT NULL COMMENT '报考书-学校招生代码',
              `enguide_maj_bkcode` varchar(20) NOT NULL COMMENT '报考书-专业招生代码',
              `enguide_schid` int(10) DEFAULT NULL COMMENT '报考书-学校id',
              `enguide_schname` varchar(20) NOT NULL COMMENT '报考书-学校名称',
              `enguide_majid` int(10) DEFAULT NULL COMMENT '报考书-专业id（弃用）',
              `enguide_majcode` varchar(20) DEFAULT NULL COMMENT '报考书-专业或者一级学科（萃言自定义编码）',
              `enguide_majname` text NOT NULL COMMENT '报考书-专业名称（报考书原始名）',
              `enguide_subjectid` int(10) DEFAULT NULL COMMENT '报考书-科类编码  299,298,300',
              `enguide_subjectname` varchar(10) DEFAULT NULL COMMENT '报考书_科类名称',
              `sm_id` int(10) DEFAULT NULL COMMENT '报考书-学校专业id',
              `hmschmajfra_id` int(10) DEFAULT NULL COMMENT '历年分数-id（当多个子类分数时id不填）',
              `major` text COMMENT '历年分数-专业名称（分数表专业原始名称）（当多个子类分数时名称用名称|拼接）',
              `subjectid` int(10) DEFAULT NULL COMMENT '历年分数_科类编码 299,298,300',
              `subjectname` varchar(10) NOT NULL COMMENT '历年分数_科类名称',
              `years` varchar(10) NOT NULL COMMENT '历年分数-年份',
              `maxfra` double(6,1) DEFAULT NULL COMMENT '历年分数-最高分',
              `minfra` double(6,1) DEFAULT NULL COMMENT '历年分数-最低分',
              `avgfra` double(6,1) DEFAULT NULL COMMENT '历年分数-平均分',
              `enguide_majstate` int(2) DEFAULT NULL COMMENT '报考书历年分数，专业名称标识（\r\n1.名称一对一\r\n2.大小类转换（跟到学校）（书名无括号）\r\n3.ES查询\r\n4. 分名含于书名（精准）（分数有括号：包含多括号、内嵌括号）\r\n5.中外合作办学\r\n6.只招少数民族\r\n10.分名含于书名（精准）（分名无括号、书名单括号（包含内嵌括号））\r\n11\r\n12\r\n',
              `sub_sumnums` int(8) DEFAULT NULL COMMENT '历年分数-科类下考生总人数',
              `minrank` int(8) DEFAULT NULL COMMENT '历年分数-最低分当年排名',
              `maxrank` int(8) DEFAULT NULL COMMENT '历年分数-最高分当年排名',
              `avgrank` int(8) DEFAULT NULL COMMENT '历年分数-平均分当年排名（暂时弃用）（程序暂时未添加）',
              `state` int(1) DEFAULT NULL COMMENT '数据异常状态',
              `batch_id` int(10) DEFAULT NULL COMMENT '历年分数-批次id',
              `batch_name` varchar(20) DEFAULT NULL COMMENT '历年分数-批次名称',
              `enguide_num` int(8) DEFAULT NULL COMMENT '报考书招生人数',
              `enguide_batch_id` int(10) DEFAULT NULL COMMENT '报考书_历年分数-批次id',
              `enguide_batch_name` varchar(20) DEFAULT NULL COMMENT '报考书_历年分数-批次名称',
              PRIMARY KEY (`id`) USING BTREE,
              KEY `index_1` (`enguide_id`) USING BTREE
            ) ENGINE=InnoDB AUTO_INCREMENT=14552 DEFAULT CHARSET=utf8mb4 COMMENT='报考书与历年分数匹配'
                            ''' % table_name_9

    do_insert_8 = 0
    do_insert_9 = 0
    conn = pymysql.connect(**config)
    cursor = conn.cursor()
    cursor.execute('''show tables like \'%s\'''' % table_name_8)
    if cursor.fetchone() is None:
        cursor.execute(sql_create_table8)
        do_insert_8 = 1
    else:
        sql_table_count = '''select count(*) from %s''' % table_name_8
        cursor.execute(sql_table_count)
        count = cursor.fetchone()[0]
        if count == 0:
            do_insert_8 = 1
    cursor.close()

    cursor = conn.cursor()
    cursor.execute('''show tables like \'%s\'''' % table_name_9)
    if cursor.fetchone() is None:
        cursor.execute(sql_create_table9)
        do_insert_9 = 1
    else:
        sql_table_count = '''select count(*) from %s''' % table_name_9
        cursor.execute(sql_table_count)
        count = cursor.fetchone()[0]
        if count == 0:
            do_insert_9 = 1
    cursor.close()


    if do_insert_8 == 0 and do_insert_9 == 0:
        print("！！！ %s%s%s 理科及文科数据均已存在，请核对!" % (parser_file_name(file_name)['province_sort'],
                                           parser_file_name(file_name)['year'], parser_file_name(file_name)['key']))
    elif do_insert_8 == 0:
        print("！！！ %s%s%s 理科数据已存在，请核对!" % (parser_file_name(file_name)['province_sort'],
                                       parser_file_name(file_name)['year'], parser_file_name(file_name)['key']))
    elif do_insert_9 == 0:
        print("！！！ %s%s%s 文科数据已存在，请核对!" % (parser_file_name(file_name)['province_sort'],
                                       parser_file_name(file_name)['year'], parser_file_name(file_name)['key']))
    else:
        print("--- 正在导入 --- %s%s%s 数据，请等待。。。" % (parser_file_name(file_name)['province_sort'],
                                        parser_file_name(file_name)['year'], parser_file_name(file_name)['key']))

        wb = openpyxl.load_workbook(file_name)  # 打开excel
        sheet = wb.worksheets[0]  # 获取第一个表
        max_row = sheet.max_row
        max_column = sheet.max_column

        column_excel = [sheet.cell(row=1, column=i).value.lower() for i in range(1, max_column + 1) if
                        sheet.cell(row=1, column=i).value is not None]
        # print(column_excel)  # excel所有字段名称

        order_reference = {
            'enguide_id': '报考书id', 'enguide_sch_bkcode': '院校代号', 'enguide_maj_bkcode': '专业代号',
            'enguide_schname': '学校名称', 'enguide_majname': '专业名称合并', 'enguide_subjectname': '科类', 'subjectname': '科类',
            'years': '年份', 'maxfra': '最高分', 'major': '专业名称去括号', 'minfra': '最低分', 'avgfra': '平均分', 'batch_name': '批次',
            'enguide_num': '招生计划'}
        sql_insert_8 = '''
        insert into {}(enguide_id,enguide_sch_bkcode,enguide_maj_bkcode,enguide_schname,enguide_majname,
                      enguide_subjectname,subjectname,years,maxfra,major,minfra,avgfra,batch_name,enguide_num,
                        enguide_schid,enguide_majid,enguide_majcode,enguide_subjectid,subjectid,batch_id)
                        values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                       '''.format(table_name_8)

        sql_insert_9 = '''
        insert into {}(enguide_id,enguide_sch_bkcode,enguide_maj_bkcode,enguide_schname,enguide_majname,
                      enguide_subjectname,subjectname,years,maxfra,major,minfra,avgfra,batch_name,enguide_num,
                        enguide_schid,enguide_majid,enguide_majcode,enguide_subjectid,subjectid,batch_id)
                        values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                       '''.format(table_name_9)

        num = 0
        unmatch = 0
        unmatch_school = set()
        unmatch_major = set()
        for i in range(2, max_row + 1):
            one_row = []
            if sheet.cell(row=i, column=1).value is not None:
                for k in order_reference.keys():
                    one_row.append(sheet.cell(row=i, column=column_excel.index(order_reference[k]) + 1).value)

                try:
                    one_row.append(sch_data[sheet.cell(row=i, column=column_excel.index(
                        order_reference['enguide_schname']) + 1).value]['id'])
                except:
                    one_row.append(None)
                    unmatch_school.add(
                        sheet.cell(row=i, column=column_excel.index(order_reference['enguide_schname']) + 1).value)

                try:
                    one_row.append(
                        major_data[
                            sheet.cell(row=i, column=column_excel.index(order_reference['major']) + 1).value][
                            'id'])
                except:
                    one_row.append(None)
                    unmatch_major.add(
                        sheet.cell(row=i, column=column_excel.index(order_reference['major']) + 1).value)

                try:
                    one_row.append(major_data[sheet.cell(row=i, column=column_excel.index(
                        order_reference['major']) + 1).value]['code'])
                except:
                    one_row.append(None)

                one_row.append(
                    subject[
                        sheet.cell(row=i, column=column_excel.index(order_reference['enguide_subjectname']) + 1).value])
                one_row.append(
                    subject[
                        sheet.cell(row=i, column=column_excel.index(order_reference['enguide_subjectname']) + 1).value])
                # 按一、二、三匹配批次
                for k in batch_data:
                    if k[:2] == province_sort and re.search("[一二三]", k).group() == re.search("[一二三]",
                                                                                             sheet.cell(row=i,
                                                                                                        column=column_excel.index(
                                                                                                            order_reference[
                                                                                                                'batch_name']) + 1).value).group():
                        one_row.append(batch_data[k]['id'])
                        break


            else:
                continue

            cursor = conn.cursor()
            for v in one_row[15:]:
                if v is None:
                    unmatch = 1
                    # print("问题数据：", i, one_row, sep=' ')
                    break
                else:
                    if one_row[-2] == 298:
                        # 导入数据库，不成功则打印
                        try:
                            cursor.execute(sql_insert_8, (
                                one_row[0], one_row[1], one_row[2], one_row[3], one_row[4], one_row[5], one_row[6],
                                one_row[7], one_row[8], one_row[9], one_row[10], one_row[11], one_row[12], one_row[13],
                                one_row[14],
                                one_row[15], one_row[16], one_row[17], one_row[18], one_row[19]))
                            num += 1
                        except:
                            print("xxx 导入失败 xxx：", one_row, sep=' ')
                    else:
                        try:
                            cursor.execute(sql_insert_9, (
                                one_row[0], one_row[1], one_row[2], one_row[3], one_row[4], one_row[5], one_row[6],
                                one_row[7], one_row[8], one_row[9], one_row[10], one_row[11], one_row[12], one_row[13],
                                one_row[14],
                                one_row[15], one_row[16], one_row[17], one_row[18], one_row[19]))
                            num += 1
                        except:
                            print("xxx 导入失败 xxx：", one_row, sep=' ')




        cursor.close()
        if unmatch == 1:
            unmatch_data = {}
            unmatch_data['学校'] = unmatch_school
            unmatch_data['专业'] = unmatch_major
            for k in list(unmatch_data.keys()):
                if unmatch_data[k] == set():
                    del unmatch_data[k]
            print("xxx 导入失败 xxx 基础表中存在缺失项：", unmatch_data, sep='')
        else:
            conn.commit()
            print("vvv 导入成功 vvv %s%s%s 共计导入%s条，请检查。" % (
                parser_file_name(file_name)['province_sort'], parser_file_name(file_name)['year'],
                parser_file_name(file_name)['key'], num))
        conn.close()


def to_guidefra_science():
    pass
